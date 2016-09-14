from openpyxl import load_workbook, Workbook
from openpyxl.cell import get_column_letter
import os, sys
import traceback
import parseAddress


fixture_folder = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "Fixtures")
infile = os.path.join(fixture_folder, "AddressesForJSK_2.xlsx")
wb = load_workbook(infile, read_only=True)
ws = wb.get_sheet_by_name("SampleData")

# read through each row in the basefile and perform the address parsing
# then write the output to the output sheet

number_of_rows = 0
for row in ws.rows:
    number_of_rows += 1

print(number_of_rows)


row_index = 2
results = {}
for x in ws.iter_rows('A{}:B{}'.format(row_index, number_of_rows)):
    rowID = x[0].value
    description = x[1].value
    parser = parseAddress.Parser(description)
    parsed = parser.usaddress_tag()
    results[rowID] = parsed
    row_index += 1

headers = []
for k, v in results.iteritems():
    for tup in v:
        headers.append(tup[1])
    headers = list(set(headers))

num_columns = len(headers)
num_rows = len(results)

out_wb = Workbook()
out_ws = out_wb.active

try:
    # write the headers on line 1
    for col in range(2, num_columns):
        out_ws["A1"] = "RowID"
        out_ws["{}1".format(get_column_letter(col))] = headers[col]

    # i represents the current row being written to starting with 2
    i = 2
    for k, v in results.iteritems():
        # set the rowID
        out_ws["A{}".format(i)] = k
        for tup in v:
            # match the value to the correct column header
            assignment = tup[1]
            for col in range(2, num_columns):
                if out_ws["{}1".format(get_column_letter(col))].value == assignment:
                    out_ws["{}{}".format(get_column_letter(col), i)] = tup[0]
        i += 1
except:
    exc_type, exc_value, exc_traceback = sys.exc_info()
    traceback.print_tb(exc_traceback, limit=4, file=sys.stdout)


out_wb.save(os.path.join(fixture_folder, "ParsedAddresses.xlsx"))
